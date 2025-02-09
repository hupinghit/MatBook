# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook

def load_materials_from_excel(file_path):
    """动态加载Excel材料数据及每个工作表的题头"""
    materials = {}
    headers_by_sheet = {}  # 存储每个工作表的题头
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        
        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if sheet.max_row < 1:
                continue
                
            # 读取第一行作为题头
            header_row = [cell.value for cell in sheet[1]] if sheet.max_row >=1 else []
            cleaned_headers = [str(hdr).strip() for hdr in header_row if hdr]
            headers_by_sheet[sheet_name] = cleaned_headers  # 存储当前工作表的题头
            
            materials[sheet_name] = {}  # {材料名称: {属性字典}}
            
            # 处理数据行
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue  # 跳过空行
                
                material_name = str(row[0]).strip()
                material_properties = {}
                
                # 填充属性
                for idx, header in enumerate(cleaned_headers):
                    value = row[idx] if idx < len(row) else ""
                    material_properties[header] = str(value).strip()
                    
                materials[sheet_name][material_name] = material_properties
                
        return materials, headers_by_sheet
    except Exception as e:
        messagebox.showerror(
            "数据加载失败",
            f"文件路径：{file_path}\n错误类型：{type(e).__name__}\n详细信息：{str(e)}"
        )
        return None, None

def build_gui(materials_data, headers_by_sheet):
    """构建图形界面"""
    root = tk.Tk()
    root.title("材料属性浏览器 v1.5")
    root.geometry("1200x600")
    
    # 使用现代化主题
    style = ttk.Style()
    style.theme_use("clam")
    
    # 左侧材料选择面板
    left_panel = ttk.Frame(root, padding=(10, 10), width=300)
    left_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=False)
    
    # 树状视图配置（材料选择）
    tree = ttk.Treeview(
        left_panel,
        show="tree",  # 仅显示树状视图，不显示列头
        selectmode="browse",
        height=30
    )
    tree.heading("#0", text="材料类别/名称", anchor=tk.W)
    tree.column("#0", width=200, stretch=tk.NO)
    
    # 滚动条
    vsb = ttk.Scrollbar(left_panel, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=vsb.set)
    vsb.pack(side=tk.RIGHT, fill=tk.Y)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    # 右侧材料属性面板
    right_panel = ttk.Frame(root, padding=(15,10))
    right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
    
    # 表格配置（材料属性）
    table = ttk.Treeview(
        right_panel,
        columns=("属性名称", "属性值"),
        show="headings",
        selectmode="browse",
        height=30
    )
    
    # 设置表格列头
    table.heading("属性名称", text="属性名称")
    table.heading("属性值", text="属性值")
    table.column("属性名称", width=150, anchor=tk.W)
    table.column("属性值", width=150, anchor=tk.W)
    
    # 表格滚动条
    table_vsb = ttk.Scrollbar(right_panel, orient=tk.VERTICAL, command=table.yview)
    table.configure(yscroll=table_vsb.set)
    table_vsb.pack(side=tk.RIGHT, fill=tk.Y)
    table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    # 填充左侧树状视图（材料选择）
    for sheet_name, items in materials_data.items():
        parent = tree.insert("", tk.END, text=sheet_name, open=False)
        for mat_name in items.keys():
            tree.insert(parent, tk.END, text=mat_name, tags=("data_row",))
    
    # 标签样式配置
    tree.tag_configure("data_row", background="#f0f0f0")
    
    # 绑定树状视图选择事件
    def on_tree_select(event):
        selected_item = tree.selection()[0]
        item_text = tree.item(selected_item, "text")
        
        # 查找材料属性
        found = False
        for sheet_name, items in materials_data.items():
            if item_text in items:
                material_properties = items[item_text]
                found = True
                break
        else:
            material_properties = {}
        
        # 更新右侧表格
        table.delete(*table.get_children())
        
        # 动态填充表格数据
        for header, value in material_properties.items():
            table.insert("", tk.END, values=(header, value))
    
    tree.bind("<<TreeviewSelect>>", on_tree_select)
    
    return root

if __name__ == "__main__":
    FILE_PATH = "materials.xlsx"
    
    # 加载数据
    materials_data, headers_by_sheet = load_materials_from_excel(FILE_PATH)
    if not materials_data or not headers_by_sheet:
        exit()
    
    # 启动GUI
    app = build_gui(materials_data, headers_by_sheet)
    app.mainloop()